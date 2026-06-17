"""
Vercel serverless API for UNC Research Graph.
Loads graph.json at cold-start and serves matching entirely in-memory.
No DuckDB, no API keys, no external calls at request time.
"""
from __future__ import annotations
import json
import math
import re
import os
import sys
import secrets
from pathlib import Path
from functools import lru_cache
from http.server import BaseHTTPRequestHandler
from urllib.parse import urlparse, parse_qs, unquote


def _log(*a):
    """Server-side diagnostics (captured in Vercel logs); never sent to clients."""
    try:
        print(*a, file=sys.stderr)
    except Exception:
        pass

# ── data loading ────────────────────────────────────────────────────────────

GRAPH_PATH = Path(__file__).parent.parent / "frontend" / "graph.json"

@lru_cache(maxsize=1)
def _load_graph() -> dict:
    # Graceful cold-start: a missing or malformed graph.json must degrade to an
    # empty graph (HTTP 200 "degraded"), never crash the function with a 500.
    if not GRAPH_PATH.exists():
        return {"meta": {"built_at": None, "n_companies": 0, "n_units_with_data": 0}, "units": [], "companies": [], "_empty": True}
    try:
        with open(GRAPH_PATH) as f:
            return json.load(f)
    except (json.JSONDecodeError, OSError) as exc:
        # Log the detail server-side, but never leak paths/exceptions to clients.
        _log("graph load failed:", repr(exc))
        return {"meta": {"built_at": None, "error": "graph unavailable"}, "units": [], "companies": [], "_empty": True}

def _graph():
    return _load_graph()

PARTNERSHIPS_PATH = Path(__file__).parent.parent / "frontend" / "partnerships.json"

@lru_cache(maxsize=1)
def _load_partnerships() -> dict:
    # Same graceful cold-start contract as the graph: missing/malformed
    # partnerships.json degrades to an empty inventory, never a 500.
    empty = {"meta": {"built_at": None, "n_units": 0, "n_faculty": 0, "n_partnerships": 0},
             "units": [], "faculty": [], "partnerships": [], "_empty": True}
    if not PARTNERSHIPS_PATH.exists():
        return empty
    try:
        with open(PARTNERSHIPS_PATH) as f:
            return json.load(f)
    except (json.JSONDecodeError, OSError) as exc:
        _log("partnerships load failed:", repr(exc))
        return {**empty, "meta": {"error": "inventory unavailable"}}

def _partnerships():
    # When a writable store (Vercel KV) is configured, prefer its live copy so
    # edits made through the write API are reflected in reads across instances.
    if _kv_enabled():
        kv = _kv_get()
        if kv is not None:
            return kv
    return _load_partnerships()

# ── editable write layer ──────────────────────────────────────────────────────
# Edits persist to a writable store, chosen at runtime:
#   • Vercel KV (env KV_REST_API_URL + KV_REST_API_TOKEN) — works on the deployed
#     serverless site; or
#   • frontend/partnerships.json on a writable filesystem (local dev).
# Writes require the EDIT_TOKEN (when configured) and a same-origin request.

MAX_WRITE_BYTES = 200_000  # reject oversized payloads

_EDITABLE_UNIT_FIELDS = {"parent_unit_id", "unit_name", "unit_type", "description",
                         "focus_areas", "disciplines", "faculty_count", "student_count",
                         "website_url", "research_by", "date_of_research", "notes"}
_EDITABLE_PARTNERSHIP_FIELDS = {"unit_id", "area", "company_name", "description", "status",
                                "start_date", "end_date", "renewal_date", "recurring",
                                "funding_value", "funding_type", "unc_poc", "company_poc",
                                "source_url", "verification_tier", "verification_notes",
                                "research_by", "date_of_research"}
_URL_FIELDS = {"website_url", "source_url"}

# ── input sanitisation ────────────────────────────────────────────────────────
# Every value that lands in the inventory is forced to a safe scalar: control
# characters stripped, strings length-capped, nested structures rejected. This
# bounds stored size and keeps non-scalar junk out of the JSON store regardless
# of what a client sends.
MAX_STR = 20_000                       # per-field character cap
MAX_FIELDS = 60                        # reject payloads with absurd key counts
_CTRL = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]")
_ID_OK = re.compile(r"[^A-Za-z0-9:_./-]")


def _clean_scalar(v):
    """Coerce a write value to a safe scalar; drop nested structures."""
    if v is None or isinstance(v, bool) or isinstance(v, (int, float)):
        return v
    if isinstance(v, str):
        s = _CTRL.sub("", v).strip()
        return s[:MAX_STR]
    return None  # lists/dicts are not valid in any editable field


def _clean_id(v) -> str:
    """Restrict client-supplied identifiers to a safe charset (length-capped)."""
    s = _CTRL.sub("", str(v or "")).strip()
    return _ID_OK.sub("", s)[:200]


_ID_PAYLOAD_FIELDS = {"unit_id", "parent_unit_id"}


def _clean_payload(payload: dict) -> dict:
    out = {}
    for k, v in payload.items():
        k = str(k)[:120]
        out[k] = _clean_id(v) if k in _ID_PAYLOAD_FIELDS else _clean_scalar(v)
    return out


# ── authn / request hardening ─────────────────────────────────────────────────

def _edit_token() -> str:
    return os.environ.get("EDIT_TOKEN", "")


def _check_auth(headers: dict) -> bool:
    """Writes are open when no EDIT_TOKEN is set (local dev); otherwise a matching
    X-Edit-Token (or Authorization: Bearer) header is required."""
    token = _edit_token()
    if not token:
        return True
    supplied = (headers or {}).get("x-edit-token", "")
    auth = (headers or {}).get("authorization", "")
    if not supplied and auth.lower().startswith("bearer "):
        supplied = auth[7:]
    return bool(supplied) and secrets.compare_digest(supplied, token)


def _same_origin(headers: dict) -> bool:
    """CSRF guard: if a browser Origin is present on a write, its host must match
    the request Host. Non-browser clients (no Origin) pass and rely on the token."""
    h = headers or {}
    origin = h.get("origin", "")
    if not origin:
        return True
    host = h.get("host", "")
    try:
        return urlparse(origin).netloc == host
    except ValueError:
        return False


def _safe_url(u):
    s = (u or "").strip()
    if s and not re.match(r"^https?://", s, re.I):
        return ""  # drop javascript:/data:/other script-bearing schemes
    return s


# ── persistence (Vercel KV when configured, else local file) ──────────────────

_KV_URL = os.environ.get("KV_REST_API_URL", "").rstrip("/")
_KV_TOKEN = os.environ.get("KV_REST_API_TOKEN", "")
_KV_KEY = "iig:inventory"


def _kv_enabled() -> bool:
    return bool(_KV_URL and _KV_TOKEN)


def _kv_get():
    import urllib.request
    req = urllib.request.Request(f"{_KV_URL}/get/{_KV_KEY}",
                                 headers={"Authorization": f"Bearer {_KV_TOKEN}"})
    try:
        with urllib.request.urlopen(req, timeout=6) as r:
            val = json.loads(r.read()).get("result")
        return json.loads(val) if val else None
    except Exception:
        return None


def _kv_set(data: dict) -> None:
    import urllib.request
    body = json.dumps(data, ensure_ascii=False).encode()
    req = urllib.request.Request(f"{_KV_URL}/set/{_KV_KEY}", data=body, method="POST",
                                 headers={"Authorization": f"Bearer {_KV_TOKEN}",
                                          "Content-Type": "application/octet-stream"})
    urllib.request.urlopen(req, timeout=6).read()


def _read_inventory() -> dict:
    """Fresh read for read-modify-write: KV if configured (seeded from the bundled
    file the first time), else the local file."""
    if _kv_enabled():
        kv = _kv_get()
        if kv is not None:
            return kv
    with open(PARTNERSHIPS_PATH) as f:
        return json.load(f)


def _save_inventory(data: dict) -> None:
    if _kv_enabled():
        _kv_set(data)
    else:
        with open(PARTNERSHIPS_PATH, "w") as f:
            json.dump(data, f, ensure_ascii=False)
    _load_partnerships.cache_clear()  # so subsequent GETs see the edit


def _slugify(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", (s or "").lower()).strip("_") or secrets.token_hex(3)


def _recount_units(data: dict) -> None:
    """Keep each unit's partnership_count in sync with the partnerships array."""
    counts: dict = {}
    for p in data.get("partnerships", []):
        uid = p.get("unit_id")
        if uid:
            counts[uid] = counts.get(uid, 0) + 1
    for u in data.get("units", []):
        u["partnership_count"] = counts.get(u["unit_id"], 0)
    meta = data.setdefault("meta", {})
    meta["n_units"] = len(data.get("units", []))
    meta["n_partnerships"] = len(data.get("partnerships", []))

# ── normalisation helpers ────────────────────────────────────────────────────

_STRIP = re.compile(r"\b(inc|llc|corp|co|ltd|plc|lp|incorporated|corporation|limited|company)\b\.?", re.I)
_PUNCT = re.compile(r"[^\w\s&]")
_WS    = re.compile(r"\s+")

def norm_name(s: str) -> str:
    s = (s or "").lower()
    s = _STRIP.sub(" ", s)
    s = _PUNCT.sub(" ", s)
    return _WS.sub(" ", s).strip()

STOPWORDS = set("the a an and or of in to for with on at by from is are was inc llc corp co ltd company group holdings international plc the study research".split())

def tokenize(s: str) -> set[str]:
    tokens = re.sub(r"[^\w\s]", " ", (s or "").lower()).split()
    return {t for t in tokens if len(t) >= 3 and t not in STOPWORDS}

# ── matching ─────────────────────────────────────────────────────────────────

def match_company(query: str, graph: dict) -> dict | None:
    q = norm_name(query)
    if len(q) < 2:
        return None
    companies = graph.get("companies", [])
    # exact
    for c in companies:
        if norm_name(c["name"]) == q:
            return c
    # substring
    subs = [c for c in companies if norm_name(c["name"]) in q or q in norm_name(c["name"])]
    if subs:
        return max(subs, key=lambda c: c.get("total_edges", 0))
    return None

def match_topical(query: str, graph: dict, top_n: int = 6) -> list[dict]:
    q = tokenize(query)
    if not q:
        return []
    results = []
    for u in graph.get("units", []):
        kw = u.get("keywords") or []
        if not kw:
            continue
        kw_set = set(kw)
        hits = [t for t in q if t in kw_set]
        if not hits:
            continue
        score = len(hits) / math.sqrt(len(q) * len(kw))
        results.append({"unit": u, "score": score, "hits": hits})
    results.sort(key=lambda x: -x["score"])
    return results[:top_n]

# ── response helpers ──────────────────────────────────────────────────────────

def json_response(data, status=200):
    body = json.dumps(data, ensure_ascii=False).encode()
    return status, {"Content-Type": "application/json"}, body

def cors_headers():
    # GET (public read API) stays open; writes are enforced server-side via the
    # edit token + same-origin check, not via CORS.
    return {
        "Access-Control-Allow-Origin": "*",
        "Access-Control-Allow-Methods": "GET, POST, PUT, DELETE, OPTIONS",
        "Access-Control-Allow-Headers": "Content-Type, X-Edit-Token, Authorization",
        "Access-Control-Max-Age": "86400",
        "X-Content-Type-Options": "nosniff",
        "Referrer-Policy": "strict-origin-when-cross-origin",
        "X-Frame-Options": "DENY",
        # JSON/exports are never a document context — lock down any accidental render.
        "Content-Security-Policy": "default-src 'none'; frame-ancestors 'none'; sandbox",
        # Belt-and-suspenders HTTPS pinning (Vercel is HTTPS-only).
        "Strict-Transport-Security": "max-age=63072000; includeSubDomains; preload",
    }

# ── partnership inventory helpers ─────────────────────────────────────────────

def _cors_json(data, status=200):
    s, headers, body = json_response(data, status)
    return s, {**headers, **cors_headers()}, body

def _unit_tree(units: list) -> list:
    """Nest units into schools -> departments -> labs/centers via parent_unit_id."""
    by_id = {u["unit_id"]: {**u, "children": []} for u in units}
    roots = []
    for u in by_id.values():
        parent = u.get("parent_unit_id")
        if parent and parent in by_id and parent != u["unit_id"]:
            by_id[parent]["children"].append(u)
        else:
            roots.append(u)
    return roots

def _filter_partnerships(rows: list, qs: dict) -> list:
    """Apply optional ?area=&status=&tier=&unit_id= filters."""
    def first(key):
        v = qs.get(key)
        return v[0] if v else None
    out = rows
    for key, field in (("area", "area"), ("status", "status"),
                       ("tier", "verification_tier"), ("unit_id", "unit_id")):
        val = first(key)
        if val:
            out = [r for r in out if r.get(field) == val]
    return out

_XLSX_UNIT_COLS = ["unit_id", "parent_unit_id", "unit_name", "unit_type", "description",
                   "focus_areas", "disciplines", "faculty_count", "student_count",
                   "website_url", "research_by", "date_of_research", "notes"]
_XLSX_PARTNERSHIP_COLS = ["unit_name", "unit_type", "faculty_name", "area", "company_name",
                          "description", "status", "start_date", "end_date", "recurring",
                          "funding_value", "funding_type", "unc_poc", "company_poc",
                          "source_url", "verification_tier", "verification_notes",
                          "research_by", "date_of_research"]
_XLSX_FACULTY_COLS = ["full_name", "title", "unit_name", "profile_url", "openalex_id",
                      "research_by", "date_of_research"]
_XLSX_TIER_FILL = {"Verified": "C6EFCE", "Reported": "FFEB9C", "Inferred": "D9D9D9"}

def _build_inventory_xlsx(data: dict) -> bytes:
    """Generate UNC_Partnership_Inventory.xlsx in-memory from partnerships.json."""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    def sheet(ws, columns, rows):
        ws.append(columns)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1D1D1F")
        tier_idx = columns.index("verification_tier") + 1 if "verification_tier" in columns else None
        for r in rows:
            ws.append([r.get(c) for c in columns])
            if tier_idx:
                color = _XLSX_TIER_FILL.get(r.get("verification_tier"))
                if color:
                    ws.cell(row=ws.max_row, column=tier_idx).fill = PatternFill("solid", fgColor=color)
        ws.freeze_panes = "A2"

    wb = Workbook()
    sheet(wb.active, _XLSX_UNIT_COLS, data.get("units", []))
    wb.active.title = "UNC_Units"
    sheet(wb.create_sheet("Partnerships"), _XLSX_PARTNERSHIP_COLS, data.get("partnerships", []))
    sheet(wb.create_sheet("Faculty"), _XLSX_FACULTY_COLS, data.get("faculty", []))
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── route handler ─────────────────────────────────────────────────────────────

def handle_write(method: str, path: str, body: bytes, headers: dict) -> tuple[int, dict, bytes]:
    """POST/PUT/DELETE for the editable units & partnerships inventory."""
    if not _same_origin(headers):
        return _cors_json({"error": "bad_origin", "message": "Cross-origin writes are not allowed."}, 403)
    if not _check_auth(headers):
        return _cors_json({"error": "unauthorized", "message": "A valid edit token is required to make changes."}, 401)
    if len(body) > MAX_WRITE_BYTES:
        return _cors_json({"error": "payload_too_large", "message": "Request body too large."}, 413)

    try:
        payload = json.loads(body or b"{}")
    except (json.JSONDecodeError, ValueError):
        return _cors_json({"error": "invalid JSON body"}, 400)
    if not isinstance(payload, dict):
        return _cors_json({"error": "JSON body must be an object"}, 400)
    if len(payload) > MAX_FIELDS:
        return _cors_json({"error": "too_many_fields", "message": "Request has too many fields."}, 400)
    # Force every value to a safe scalar (control chars stripped, length-capped,
    # nested structures dropped), then sanitize URL-bearing fields.
    payload = _clean_payload(payload)
    for f in _URL_FIELDS:
        if f in payload:
            payload[f] = _safe_url(payload[f])

    try:
        data = _read_inventory()
    except (OSError, json.JSONDecodeError) as exc:
        _log("inventory read failed:", repr(exc))
        return _cors_json({"error": "inventory unavailable"}, 503)

    units = data.setdefault("units", [])
    parts = data.setdefault("partnerships", [])

    def commit(result, status=200):
        _recount_units(data)
        try:
            _save_inventory(data)
        except OSError:
            # Read-only filesystem (no KV configured) — surface clearly, don't 500.
            return _cors_json({"error": "read_only",
                               "message": "This deployment is read-only; configure Vercel KV (or run locally) to persist edits."}, 503)
        except Exception as exc:  # KV/network failure
            _log("persist failed:", repr(exc))
            return _cors_json({"error": "persist_failed", "message": "Edit could not be saved. Please retry."}, 502)
        return _cors_json(result, status)

    # ── units ────────────────────────────────────────────────────────────────
    if re.match(r"^/?(?:api/)?units/?$", path) and method == "POST":
        name = (payload.get("unit_name") or "").strip()
        if not name:
            return _cors_json({"error": "unit_name is required"}, 400)
        parent = _clean_id(payload.get("parent_unit_id")) or "unc:root"
        uid = _clean_id(payload.get("unit_id")) or f"{parent}:{_slugify(name)}"
        if any(u["unit_id"] == uid for u in units):
            uid = f"{uid}_{secrets.token_hex(2)}"
        unit = {"unit_id": uid, "parent_unit_id": parent, "unit_name": name,
                "unit_type": payload.get("unit_type") or "Department",
                "description": payload.get("description") or "", "focus_areas": payload.get("focus_areas") or "",
                "disciplines": payload.get("disciplines") or "", "faculty_count": payload.get("faculty_count"),
                "student_count": payload.get("student_count"), "website_url": payload.get("website_url") or "",
                "research_by": payload.get("research_by") or "", "date_of_research": payload.get("date_of_research") or "",
                "notes": payload.get("notes") or "", "partnership_count": 0, "top_companies": []}
        units.append(unit)
        return commit(unit, 201)

    m = re.match(r"^/?(?:api/)?units/(.+)$", path)
    if m and method in ("PUT", "DELETE"):
        uid = unquote(m.group(1))
        unit = next((u for u in units if u["unit_id"] == uid), None)
        if not unit:
            return _cors_json({"error": f"unit '{uid}' not found"}, 404)
        if method == "DELETE":
            data["units"] = [u for u in units if u["unit_id"] != uid]
            return commit({"deleted": uid})
        for k, v in payload.items():
            if k in _EDITABLE_UNIT_FIELDS:
                unit[k] = v
        return commit(unit)

    # ── partnerships ───────────────────────────────────────────────────────────
    if re.match(r"^/?(?:api/)?partnerships/?$", path) and method == "POST":
        uid = _clean_id(payload.get("unit_id"))
        unit = next((u for u in units if u["unit_id"] == uid), None)
        pid = "man_" + secrets.token_hex(5)
        row = {"partnership_id": pid, "unit_id": uid,
               "unit_name": unit["unit_name"] if unit else payload.get("unit_name", ""),
               "unit_type": unit.get("unit_type") if unit else "", "faculty_id": None, "faculty_name": "",
               "area": payload.get("area") or "", "company_name": payload.get("company_name") or "",
               "description": payload.get("description") or "", "status": payload.get("status") or "In Discussion",
               "start_date": payload.get("start_date") or "", "end_date": payload.get("end_date") or "",
               "renewal_date": payload.get("renewal_date") or "", "recurring": payload.get("recurring") or "",
               "funding_value": payload.get("funding_value"), "funding_type": payload.get("funding_type") or "",
               "unc_poc": payload.get("unc_poc") or "", "company_poc": payload.get("company_poc") or "",
               "source_url": payload.get("source_url") or "",
               "verification_tier": payload.get("verification_tier") or "Inferred",
               "verification_notes": payload.get("verification_notes") or "",
               "research_by": payload.get("research_by") or "", "date_of_research": payload.get("date_of_research") or ""}
        parts.append(row)
        return commit(row, 201)

    m = re.match(r"^/?(?:api/)?partnerships/(.+)$", path)
    if m and method in ("PUT", "DELETE"):
        pid = unquote(m.group(1))
        row = next((p for p in parts if p.get("partnership_id") == pid), None)
        if not row:
            return _cors_json({"error": f"partnership '{pid}' not found"}, 404)
        if method == "DELETE":
            data["partnerships"] = [p for p in parts if p.get("partnership_id") != pid]
            return commit({"deleted": pid})
        for k, v in payload.items():
            if k in _EDITABLE_PARTNERSHIP_FIELDS:
                row[k] = v
        # keep denormalized unit name in sync if the unit changed
        if "unit_id" in payload:
            u = next((u for u in units if u["unit_id"] == payload["unit_id"]), None)
            if u:
                row["unit_name"], row["unit_type"] = u["unit_name"], u.get("unit_type")
        return commit(row)

    return _cors_json({"error": "Not found", "path": path, "method": method}, 404)


def handle(method: str, path: str, qs: dict, body: bytes = b"", headers: dict | None = None) -> tuple[int, dict, bytes]:
    graph = _graph()

    if method == "OPTIONS":
        return 204, cors_headers(), b""

    if method in ("POST", "PUT", "DELETE"):
        return handle_write(method, path, body, headers or {})

    # /health
    if path in ("/health", "/api/health"):
        meta = graph.get("meta", {})
        if graph.get("_empty"):
            # Graph not built / unreadable — report degraded but stay HTTP 200 so
            # the frontend health check doesn't treat it as a hard failure.
            status, headers, body = json_response({"status": "degraded", "reason": "graph not built", "service": "unc-research-graph"})
            return status, {**headers, **cors_headers()}, body
        status, headers, body = json_response({"status": "ok", "service": "unc-research-graph", "n_companies": meta.get("n_companies"), "n_units": meta.get("n_units_with_data"), "built_at": meta.get("built_at")})
        return status, {**headers, **cors_headers()}, body

    # /stats
    if path in ("/stats", "/api/stats"):
        meta = graph.get("meta", {})
        status, headers, body = json_response({"meta": meta, "n_units": len(graph.get("units", [])), "n_companies": len(graph.get("companies", []))})
        return status, {**headers, **cors_headers()}, body

    # /freshness — the build/coverage stamp referenced by the mission
    if path in ("/freshness", "/api/freshness"):
        meta = graph.get("meta", {})
        status, headers, body = json_response({
            "built_at": meta.get("built_at"),
            "n_companies": meta.get("n_companies", 0),
            "n_units_with_data": meta.get("n_units_with_data", 0),
            "counts": meta.get("counts", {}),
        })
        return status, {**headers, **cors_headers()}, body

    # /graph — full node/link graph for the network view (API-driven, so the
    # frontend never reads the static graph.json file directly).
    if path in ("/graph", "/api/graph"):
        meta = graph.get("meta", {})
        status, headers, body = json_response({
            "meta": meta,
            "units": graph.get("units", []),
            "companies": graph.get("companies", []),
        })
        return status, {**headers, **cors_headers()}, body

    # /match/{company}
    m = re.match(r"^/?(?:api/)?match/(.+)$", path)
    if m:
        company_name = unquote(m.group(1))[:300]
        sector_hint = (qs.get("sector_hint") or [""])[0][:300]
        try:
            top_n = int((qs.get("top_n") or ["10"])[0])
        except (TypeError, ValueError):
            top_n = 10
        top_n = max(1, min(top_n, 50))

        co = match_company(company_name, graph)
        topical = match_topical(company_name + " " + sector_hint, graph, top_n)

        unit_by_id = {u["id"]: u for u in graph.get("units", [])}
        # Enrich each company→unit link with a human-readable unit name so the
        # frontend shows "UNC School of Medicine" instead of the raw id "unc:som".
        if co and co.get("units"):
            co = {**co, "units": [
                {**cu, "unit_name": cu.get("unit_name") or (unit_by_id.get(cu.get("unit_id"), {}).get("name")) or cu.get("unit_id")}
                for cu in co["units"]
            ]}
        topical_out = []
        for item in topical:
            u = item["unit"]
            topical_out.append({"unit_id": u["id"], "unit_name": u.get("name"), "score": round(item["score"], 4), "hits": item["hits"], "footprint": u.get("footprint", {})})

        result = {
            "query": company_name,
            "company": co,
            "topical_matches": topical_out,
        }
        status, headers, body = json_response(result)
        return status, {**headers, **cors_headers()}, body

    # /unit/{id}/profile
    m = re.match(r"^/?(?:api/)?unit/(.+)/profile$", path)
    if m:
        unit_id = unquote(m.group(1))
        unit_by_id = {u["id"]: u for u in graph.get("units", [])}
        u = unit_by_id.get(unit_id)
        if not u:
            status, headers, body = json_response({"error": f"Unit '{unit_id}' not found"}, 404)
            return status, {**headers, **cors_headers()}, body
        status, headers, body = json_response(u)
        return status, {**headers, **cors_headers()}, body

    # ── partnership inventory endpoints ──────────────────────────────────────
    pg = _partnerships()

    if path in ("/units", "/api/units"):
        return _cors_json(pg.get("units", []))
    if path in ("/units/tree", "/api/units/tree"):
        return _cors_json(_unit_tree(pg.get("units", [])))

    if path in ("/faculty", "/api/faculty"):
        return _cors_json(pg.get("faculty", []))
    m = re.match(r"^/?(?:api/)?faculty/(.+)$", path)
    if m:
        fid = unquote(m.group(1))
        fac = next((f for f in pg.get("faculty", []) if f.get("faculty_id") == fid), None)
        if not fac:
            return _cors_json({"error": f"Faculty '{fid}' not found"}, 404)
        ps = [p for p in pg.get("partnerships", []) if p.get("faculty_id") == fid]
        return _cors_json({**fac, "partnerships": ps})

    # export must precede the bare /partnerships match
    if path in ("/partnerships/export", "/api/partnerships/export"):
        if pg.get("_empty"):
            return _cors_json({"error": "inventory not built"}, 503)
        xlsx = _build_inventory_xlsx(pg)
        headers = {
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "Content-Disposition": "attachment; filename=UNC_Partnership_Inventory.xlsx",
            **cors_headers(),
        }
        return 200, headers, xlsx

    if path in ("/partnerships", "/api/partnerships"):
        rows = _filter_partnerships(pg.get("partnerships", []), qs)
        return _cors_json({"count": len(rows), "partnerships": rows})

    # /unit/{id}/partnerships, /unit/{id}/faculty, /unit/{id}
    m = re.match(r"^/?(?:api/)?unit/(.+)/partnerships$", path)
    if m:
        uid = unquote(m.group(1))
        rows = [p for p in pg.get("partnerships", []) if p.get("unit_id") == uid]
        return _cors_json({"unit_id": uid, "count": len(rows), "partnerships": rows})
    m = re.match(r"^/?(?:api/)?unit/(.+)/faculty$", path)
    if m:
        uid = unquote(m.group(1))
        fac = [f for f in pg.get("faculty", []) if f.get("unit_id") == uid]
        return _cors_json({"unit_id": uid, "count": len(fac), "faculty": fac})
    m = re.match(r"^/?(?:api/)?unit/([^/]+)$", path)
    if m:
        uid = unquote(m.group(1))
        unit = next((u for u in pg.get("units", []) if u.get("unit_id") == uid), None)
        if not unit:
            return _cors_json({"error": f"Unit '{uid}' not found"}, 404)
        return _cors_json(unit)

    # 404
    status, headers, body = json_response({"error": "Not found", "path": path}, 404)
    return status, {**headers, **cors_headers()}, body


# ── Vercel handler ─────────────────────────────────────────────────────────────

class handler(BaseHTTPRequestHandler):
    def do_OPTIONS(self):
        self._respond("OPTIONS")

    def do_GET(self):
        self._respond("GET")

    def do_POST(self):
        self._respond("POST")

    def do_PUT(self):
        self._respond("PUT")

    def do_DELETE(self):
        self._respond("DELETE")

    def _respond(self, method):
        parsed = urlparse(self.path)
        qs = parse_qs(parsed.query)
        length = int(self.headers.get("Content-Length") or 0)
        req_body = self.rfile.read(length) if length else b""
        req_headers = {k.lower(): v for k, v in self.headers.items()}
        status, headers, body = handle(method, parsed.path, qs, req_body, req_headers)
        self.send_response(status)
        for k, v in headers.items():
            self.send_header(k, v)
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def log_message(self, *args):
        pass
