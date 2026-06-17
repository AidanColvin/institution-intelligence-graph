"""
Export the UNC partnership inventory.

Reads unc_units, unc_faculty, unc_partnerships from DuckDB and writes:
  - UNC_Partnership_Inventory.xlsx  (3 formatted sheets, for humans)
  - frontend/partnerships.json      (compact, for the live app + API)

Usage:
  python scripts/export_partnerships.py [--db graph.db]
      [--xlsx UNC_Partnership_Inventory.xlsx] [--json frontend/partnerships.json]
      [--built-at ISO]
"""
from __future__ import annotations

import argparse
import datetime as dt
import json
import logging
import os
import sys
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

sys.path.insert(0, str(Path(__file__).parent.parent))

from backend.graph import store                                      # noqa: E402
from backend.partnerships import common                              # noqa: E402

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(name)s %(message)s",
)
logger = logging.getLogger("export_partnerships")

MAX_COL_WIDTH = 60
TIER_FILLS = {
    common.TIER_VERIFIED: PatternFill("solid", fgColor="C6EFCE"),   # green
    common.TIER_REPORTED: PatternFill("solid", fgColor="FFEB9C"),   # yellow
    common.TIER_INFERRED: PatternFill("solid", fgColor="D9D9D9"),   # gray
}
HEADER_FONT = Font(bold=True)
HEADER_FILL = PatternFill("solid", fgColor="1D1D1F")
HEADER_WHITE = Font(bold=True, color="FFFFFF")

UNIT_TYPE_ORDER = {
    common.UNIT_SCHOOL: 0, common.UNIT_DEPARTMENT: 1, common.UNIT_LAB: 2,
    common.UNIT_CENTER: 3, common.UNIT_INSTITUTE: 4,
}

UNIT_COLUMNS = ["unit_id", "parent_unit_id", "unit_name", "unit_type", "description",
                "focus_areas", "disciplines", "faculty_count", "student_count",
                "website_url", "research_by", "date_of_research", "notes"]
PARTNERSHIP_COLUMNS = ["unit_name", "unit_type", "faculty_name", "area", "company_name",
                       "description", "status", "start_date", "end_date", "recurring",
                       "funding_value", "funding_type", "unc_poc", "company_poc",
                       "source_url", "verification_tier", "verification_notes",
                       "research_by", "date_of_research"]
FACULTY_COLUMNS = ["full_name", "title", "unit_name", "profile_url", "openalex_id",
                   "research_by", "date_of_research"]


def _jsonable(value: Any) -> Any:
    """Convert DuckDB date/datetime values to ISO strings for JSON/xlsx."""
    if isinstance(value, (dt.date, dt.datetime)):
        return value.isoformat()
    return value


def _enrich(units: list[dict], faculty: list[dict],
            partnerships: list[dict]) -> list[dict]:
    """Add unit_name, unit_type and faculty_name to each partnership row."""
    unit_by_id = {u["unit_id"]: u for u in units}
    fac_by_id = {f["faculty_id"]: f for f in faculty}
    enriched: list[dict] = []
    for p in partnerships:
        unit = unit_by_id.get(p.get("unit_id"), {})
        fac = fac_by_id.get(p.get("faculty_id"), {})
        row = {k: _jsonable(v) for k, v in p.items()}
        row["unit_name"] = unit.get("unit_name")
        row["unit_type"] = unit.get("unit_type")
        row["faculty_name"] = fac.get("full_name")
        enriched.append(row)
    enriched.sort(key=lambda r: (r.get("unit_name") or "~", r.get("area") or "",
                                 r.get("company_name") or ""))
    return enriched


def _autosize(ws: Worksheet, n_cols: int) -> None:
    """Set column widths to fit content, capped at MAX_COL_WIDTH."""
    for col in range(1, n_cols + 1):
        letter = get_column_letter(col)
        widest = max((len(str(c.value)) for c in ws[letter] if c.value is not None),
                     default=10)
        ws.column_dimensions[letter].width = min(MAX_COL_WIDTH, max(10, widest + 2))


def _write_sheet(ws: Worksheet, columns: list[str], rows: list[dict],
                 generated_note: str | None = None) -> None:
    """Write a header + data rows to a sheet with freeze, bold/dark header, autosize."""
    ws.append(columns)
    for cell in ws[1]:
        cell.font = HEADER_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center")
    if generated_note:
        ws["A1"].comment = Comment(generated_note, "partnership-scan")
    tier_idx = columns.index("verification_tier") + 1 if "verification_tier" in columns else None
    for row in rows:
        ws.append([_jsonable(row.get(col)) for col in columns])
        if tier_idx is not None:
            fill = TIER_FILLS.get(row.get("verification_tier"))
            if fill is not None:
                ws.cell(row=ws.max_row, column=tier_idx).fill = fill
    ws.freeze_panes = "A2"
    _autosize(ws, len(columns))


def write_xlsx(path: Path, units: list[dict], faculty: list[dict],
               partnerships: list[dict], built_at: str) -> None:
    """Build the 3-sheet UNC_Partnership_Inventory.xlsx."""
    note = f"Generated: {built_at}"
    units_sorted = sorted(units, key=lambda u: (UNIT_TYPE_ORDER.get(u.get("unit_type"), 9),
                                                u.get("unit_name") or ""))
    fac_by_unit = {u["unit_id"]: u.get("unit_name") for u in units}
    faculty_rows = [{**f, "unit_name": fac_by_unit.get(f.get("unit_id"))} for f in faculty]
    faculty_rows.sort(key=lambda f: f.get("full_name") or "")

    wb = Workbook()
    _write_sheet(wb.active, UNIT_COLUMNS, units_sorted, note)
    wb.active.title = "UNC_Units"
    _write_sheet(wb.create_sheet("Partnerships"), PARTNERSHIP_COLUMNS, partnerships, note)
    _write_sheet(wb.create_sheet("Faculty"), FACULTY_COLUMNS, faculty_rows, note)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    logger.info("Wrote %s: %d units, %d partnerships, %d faculty",
                path, len(units), len(partnerships), len(faculty))


def write_json(path: Path, units: list[dict], faculty: list[dict],
               partnerships: list[dict], built_at: str) -> None:
    """Write the compact JSON the live app + API read (mirrors graph.json)."""
    by_unit: dict[str, list[dict]] = {}
    for p in partnerships:
        by_unit.setdefault(p.get("unit_id"), []).append(p)

    unit_out = []
    for u in units:
        ps = by_unit.get(u["unit_id"], [])
        companies = [p["company_name"] for p in ps if p.get("company_name")]
        unit_out.append({
            **{k: _jsonable(v) for k, v in u.items()},
            "partnership_count": len(ps),
            "top_companies": [c for c, _ in Counter(companies).most_common(3)],
        })

    fac_by_id_ps: dict[str, list[dict]] = {}
    for p in partnerships:
        if p.get("faculty_id"):
            fac_by_id_ps.setdefault(p["faculty_id"], []).append(p)
    unit_name = {u["unit_id"]: u.get("unit_name") for u in units}
    fac_out = []
    for f in faculty:
        ps = fac_by_id_ps.get(f["faculty_id"], [])
        companies = [p["company_name"] for p in ps if p.get("company_name")]
        fac_out.append({
            **{k: _jsonable(v) for k, v in f.items()},
            "unit_name": unit_name.get(f.get("unit_id")),
            "partnership_count": len(ps),
            "top_company": Counter(companies).most_common(1)[0][0] if companies else None,
        })

    payload = {
        "meta": {
            "built_at": built_at,
            "n_units": len(units),
            "n_faculty": len(faculty),
            "n_partnerships": len(partnerships),
            "counts_by_area": dict(Counter(p.get("area") for p in partnerships)),
            "counts_by_tier": dict(Counter(p.get("verification_tier") for p in partnerships)),
            "n_companies": len({p.get("company_name") for p in partnerships if p.get("company_name")}),
        },
        "units": unit_out,
        "faculty": fac_out,
        "partnerships": partnerships,
    }
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, separators=(",", ":")))
    logger.info("Wrote %s (%.0f KB)", path, path.stat().st_size / 1024)


def export(db_path: str | None, xlsx_path: str, json_path: str, built_at: str) -> None:
    """Read the inventory tables and write both the xlsx and the JSON."""
    if db_path:
        os.environ["GRAPH_DB_PATH"] = db_path
    units = [{k: _jsonable(v) for k, v in u.items()} for u in store.fetch_partnership_units()]
    faculty = [{k: _jsonable(v) for k, v in f.items()} for f in store.fetch_partnership_faculty()]
    partnerships = _enrich(units, faculty, store.fetch_partnerships())
    write_xlsx(Path(xlsx_path), units, faculty, partnerships, built_at)
    write_json(Path(json_path), units, faculty, partnerships, built_at)


def main() -> None:
    parser = argparse.ArgumentParser(description="Export UNC partnership inventory (xlsx + json)")
    parser.add_argument("--db", default=None, help="DuckDB path (overrides GRAPH_DB_PATH)")
    parser.add_argument("--xlsx", default="UNC_Partnership_Inventory.xlsx")
    parser.add_argument("--json", default="frontend/partnerships.json")
    parser.add_argument("--built-at", default=common.iso_now(),
                        help="ISO timestamp stamped into both outputs")
    args = parser.parse_args()
    export(args.db, args.xlsx, args.json, args.built_at)
    store.close()


if __name__ == "__main__":
    main()
